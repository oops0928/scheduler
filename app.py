import streamlit as st
import pandas as pd
import calendar
import random
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="매장 스케줄러", layout="wide", initial_sidebar_state="expanded")

WEEKDAYS_KR = ["월", "화", "수", "목", "금", "토", "일"]
OPT_WORK, OPT_OFF, OPT_LEAVE = "근무", "휴무", "연차"
CELL_OPTIONS = [OPT_WORK, OPT_OFF, OPT_LEAVE]
FONT_NAME = "맑은 고딕"


def make_dates(year, month):
    n = calendar.monthrange(year, month)[1]
    return [(d, WEEKDAYS_KR[calendar.weekday(year, month, d)]) for d in range(1, n + 1)]


def day_label(month, d, wd):
    return f"{month}/{d}({wd})"


# ==========================================
# 스케줄 생성
# ==========================================
def generate_schedule(year, month, employees, mandatory_dict, off_target_dict,
                      min_workers, close_workers_count, store_closed_days):
    days = make_dates(year, month)
    all_days = [d for d, _ in days]
    closed = set(store_closed_days)
    open_days = [d for d in all_days if d not in closed]

    sched = {e: {} for e in employees}          # emp -> {day: (code, kind)}
    off_set = {e: set() for e in employees}
    off_type = {e: {} for e in employees}
    warnings = []

    # 1. 필수 휴무 고정 (정휴일은 무시)
    for e in employees:
        for d, t in mandatory_dict[e].items():
            if d in closed:
                continue
            off_set[e].add(d)
            off_type[e][d] = t

    # 2. 남은 랜덤 휴무 일수 계산 (휴무=포함 / 연차=추가)
    remaining = {}
    for e in employees:
        counted = sum(1 for d, t in mandatory_dict[e].items()
                      if t == OPT_OFF and d not in closed)
        remaining[e] = max(0, off_target_dict[e] - counted)

    max_off = max(0, len(employees) - min_workers)
    day_off = {d: 0 for d in open_days}
    for e in employees:
        for d in off_set[e]:
            if d in open_days:
                day_off[d] += 1
    for d in open_days:
        if day_off[d] > max_off:
            warnings.append(f"⚠️ {month}/{d}: 필수 휴무가 몰려 최소 근무 인원({min_workers}명) 미달")

    # 3. 남은 휴무 공평 랜덤 배분
    order = employees[:]
    random.shuffle(order)
    for e in order:
        assigned = 0
        while assigned < remaining[e]:
            cand = [d for d in open_days if d not in off_set[e] and day_off[d] < max_off]
            if not cand:
                break
            mn = min(day_off[d] for d in cand)
            best = [d for d in cand if day_off[d] == mn]
            c = random.choice(best)
            off_set[e].add(c)
            off_type[e][c] = "랜덤"
            day_off[c] += 1
            assigned += 1
        if assigned < remaining[e]:
            warnings.append(f"⚠️ {e}: 목표 휴무 {remaining[e] - assigned}일 부족 (인원/조건이 빡빡)")

    # 4. 근무자 오픈/마감 배정
    shift_counts = {e: {'A': 0, 'B': 0, 'C': 0, 'OFF': 0, 'LEAVE': 0} for e in employees}
    for d, wd in days:
        if d in closed:
            for e in employees:
                sched[e][d] = ("정휴", "store_closed")
            continue
        is_weekend = wd in ("금", "토", "일")
        close_name = 'C' if is_weekend else 'B'

        working = [e for e in employees if d not in off_set[e]]
        for e in employees:
            if d in off_set[e]:
                if off_type[e].get(d) == OPT_LEAVE:
                    shift_counts[e]['LEAVE'] += 1
                    sched[e][d] = ("연차", "leave")
                else:
                    shift_counts[e]['OFF'] += 1
                    sched[e][d] = ("OFF", "off_req" if off_type[e].get(d) == OPT_OFF else "off_rand")

        if len(working) < min_workers:
            warnings.append(f"⚠️ {month}/{d} 근무 {len(working)}명 (최소 {min_workers}명 미달)")

        if len(working) >= close_workers_count:
            working.sort(key=lambda x: shift_counts[x]['B'] + shift_counts[x]['C'])
            pool = working[:min(len(working), close_workers_count + 2)]
            random.shuffle(pool)
            close_w = pool[:close_workers_count]
            open_w = [e for e in working if e not in close_w]
        else:
            close_w, open_w = working, []

        for e in close_w:
            shift_counts[e][close_name] += 1
            sched[e][d] = (close_name, "work")
        for e in open_w:
            shift_counts[e]['A'] += 1
            sched[e][d] = ("A", "work")

    warnings = list(dict.fromkeys(warnings))
    return sched, shift_counts, warnings


# ==========================================
# 매트릭스형 엑셀 생성
# ==========================================
def build_excel(store_name, year, month, off_count, employees, emp_info, sched):
    days = make_dates(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = "스케줄표"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    YELLOW = PatternFill("solid", fgColor="FFFF00")   # 정휴
    LIGHT = PatternFill("solid", fgColor="FFF2CC")     # 스케줄 신청(휴무/연차)
    HEADGRAY = PatternFill("solid", fgColor="F2F2F2")

    def setf(cell, bold=False, color="000000", size=10):
        cell.font = Font(name=FONT_NAME, bold=bold, color=color, size=size)
        cell.alignment = center
        cell.border = border

    DATE_C0 = 6  # F열부터 날짜

    # 상단 정보
    ws.cell(1, 1, store_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(1, 1).font = Font(name=FONT_NAME, bold=True, size=12)
    ws.cell(1, 1).alignment = center
    ws.cell(2, 1, f"{year}-{month:02d}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(2, 1).font = Font(name=FONT_NAME, bold=True, size=11)
    ws.cell(2, 1).alignment = center
    ws.cell(3, 1, "휴무일")
    ws.cell(3, 1).font = Font(name=FONT_NAME, size=10)
    ws.cell(3, 1).alignment = center
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(3, 3, off_count)
    ws.cell(3, 3).font = Font(name=FONT_NAME, size=10)
    ws.cell(3, 3).alignment = center
    for r in (1, 2, 3):
        for c in range(1, 6):
            ws.cell(r, c).border = border

    # 헤더 (4행=날짜, 5행=요일)
    HR1, HR2 = 4, 5
    ws.cell(HR1, 1, "직급"); ws.merge_cells(start_row=HR1, start_column=1, end_row=HR2, end_column=1)
    ws.cell(HR1, 2, "성명"); ws.merge_cells(start_row=HR1, start_column=2, end_row=HR2, end_column=2)
    ws.cell(HR1, 3, "연차"); ws.merge_cells(start_row=HR1, start_column=3, end_row=HR1, end_column=5)
    ws.cell(HR2, 3, "발생"); ws.cell(HR2, 4, "사용"); ws.cell(HR2, 5, "잔여")
    for c in (1, 2, 3, 4, 5):
        setf(ws.cell(HR1, c), bold=True); ws.cell(HR1, c).fill = HEADGRAY
        setf(ws.cell(HR2, c), bold=True); ws.cell(HR2, c).fill = HEADGRAY

    for i, (d, wd) in enumerate(days):
        col = DATE_C0 + i
        col_color = "0070C0" if wd == "토" else ("FF0000" if wd == "일" else "000000")
        dc = ws.cell(HR1, col, f"{month}/{d}")
        wc = ws.cell(HR2, col, wd)
        setf(dc, bold=True, color=col_color); dc.fill = HEADGRAY
        setf(wc, bold=True, color=col_color); wc.fill = HEADGRAY

    # 직원 행
    for ri, e in enumerate(employees):
        r = HR2 + 1 + ri
        info = emp_info.get(e, {})
        ws.cell(r, 1, info.get("rank", ""))
        ws.cell(r, 2, e)
        ws.cell(r, 3, info.get("gen", 0))
        ws.cell(r, 4, info.get("used", 0))
        ws.cell(r, 5, f"=C{r}-D{r}")
        for c in (1, 2, 3, 4, 5):
            setf(ws.cell(r, c))
        ws.cell(r, 2).font = Font(name=FONT_NAME, bold=True, size=10)

        for i, (d, wd) in enumerate(days):
            col = DATE_C0 + i
            code, kind = sched[e].get(d, ("", "work"))
            cell = ws.cell(r, col, code)
            setf(cell, bold=(kind in ("off_req", "leave")))
            if kind == "store_closed":
                cell.fill = YELLOW
            elif kind in ("off_req", "leave"):
                cell.fill = LIGHT

    # 정휴 컬럼 헤더 강조
    for i, (d, wd) in enumerate(days):
        if any(sched[e].get(d, ("", ""))[1] == "store_closed" for e in employees):
            col = DATE_C0 + i
            ws.cell(HR1, col).fill = YELLOW
            ws.cell(HR2, col).fill = YELLOW

    note_r = HR2 + 1 + len(employees) + 1
    ws.cell(note_r, 1,
            "굵은 글씨 = 스케줄 신청 건 / 노란색 = 정휴(매장 휴점) / A=오픈, B=평일마감, C=주말마감")
    ws.cell(note_r, 1).font = Font(name=FONT_NAME, size=9, color="808080")

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 9
    for c in ("C", "D", "E"):
        ws.column_dimensions[c].width = 6
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(DATE_C0 + i)].width = 5.5
    for r in range(HR2 + 1, HR2 + 1 + len(employees)):
        ws.row_dimensions[r].height = 26
    ws.freeze_panes = ws.cell(HR2 + 1, DATE_C0)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==========================================
# 🎨 UI
# ==========================================
st.title("📅 매장 자동 스케줄러")

with st.sidebar:
    st.header("⚙️ 기본 설정")
    store_name = st.text_input("매장 이름", "ADP 더현대 서울")
    year = st.selectbox("연도", [2025, 2026, 2027], index=1)
    month = st.selectbox("월", list(range(1, 13)), index=5)
    st.markdown("---")
    st.header("🔧 근무 조건")
    default_off = st.number_input("월 기본 휴무 일수", min_value=0, max_value=31, value=9)
    min_workers = st.number_input("하루 최소 근무 인원", min_value=1, max_value=20, value=3)
    close_workers_count = st.number_input("마감조(B/C) 필요 인원", min_value=1, max_value=20, value=2)

days = make_dates(year, month)
date_labels = [day_label(month, d, wd) for d, wd in days]
label_to_day = {day_label(month, d, wd): d for d, wd in days}

# --- 직원 정보 입력 ---
st.header("👥 직원 정보")
st.caption("직급 / 성명 / 연차 발생·사용을 입력하세요. (잔여 = 발생 − 사용, 자동 계산)")

default_info = pd.DataFrame([
    {"직급": "매니저",  "성명": "박미진", "연차발생": 5,  "연차사용": 0},
    {"직급": "부매니저", "성명": "김혜영", "연차발생": 2,  "연차사용": 0},
    {"직급": "주니어",  "성명": "황별이", "연차발생": 0,  "연차사용": 0},
    {"직급": "주니어",  "성명": "장재원", "연차발생": 15, "연차사용": 0},
    {"직급": "주니어",  "성명": "최지호", "연차발생": 2,  "연차사용": 0},
])
if "emp_info_df" not in st.session_state:
    st.session_state.emp_info_df = default_info

emp_info_df = st.data_editor(
    st.session_state.emp_info_df,
    num_rows="dynamic",
    use_container_width=True,
    key="emp_info_editor",
    column_config={
        "직급": st.column_config.TextColumn("직급"),
        "성명": st.column_config.TextColumn("성명", required=True),
        "연차발생": st.column_config.NumberColumn("연차 발생", min_value=0, step=1),
        "연차사용": st.column_config.NumberColumn("연차 사용", min_value=0, step=1),
    },
)
st.session_state.emp_info_df = emp_info_df

employees = [str(n).strip() for n in emp_info_df["성명"].tolist() if str(n).strip()]
emp_info = {}
for _, row in emp_info_df.iterrows():
    nm = str(row["성명"]).strip()
    if nm:
        emp_info[nm] = {
            "rank": str(row.get("직급", "") or ""),
            "gen": int(row.get("연차발생", 0) or 0),
            "used": int(row.get("연차사용", 0) or 0),
        }

# --- 정휴일(매장 휴점) ---
st.markdown("---")
store_closed_labels = st.multiselect(
    "🏬 정휴일 (매장 전체 휴점 — 전원 '정휴'로 표시, 개인 휴무일수에서 제외)",
    options=date_labels,
)
store_closed_days = [label_to_day[l] for l in store_closed_labels]

# 직원별 휴무 일수 (선택)
off_target_dict = {e: int(default_off) for e in employees}
with st.expander("👤 직원별 휴무 일수 따로 설정 (선택 · 파트타임 등)"):
    st.caption("비워두면 위의 '월 기본 휴무 일수'를 사용합니다.")
    for e in employees:
        off_target_dict[e] = st.number_input(
            e, min_value=0, max_value=31, value=int(default_off), key=f"off_{e}"
        )

# --- 필수 휴무일 입력표 ---
st.markdown("---")
st.header("🏝️ 꼭 쉬어야 하는 날 입력")
st.info(
    f"반드시 쉬어야 하는 날만 선택하세요. `{OPT_OFF}`=기본 휴무에 포함 / "
    f"`{OPT_LEAVE}`=기본 휴무에 추가(연차). 나머지는 자동으로 공평하게 배분됩니다."
)

sig = (tuple(employees), year, month)
if "mand_sig" not in st.session_state or st.session_state.mand_sig != sig:
    st.session_state.mand_df = pd.DataFrame(OPT_WORK, index=date_labels, columns=employees)
    st.session_state.mand_sig = sig

col_cfg = {}
for e in employees:
    cnt = int((st.session_state.mand_df[e] != OPT_WORK).sum()) if e in st.session_state.mand_df else 0
    col_cfg[e] = st.column_config.SelectboxColumn(
        f"👤 {e} (필수 {cnt}일)", options=CELL_OPTIONS, required=True, width="small"
    )

generate_clicked = st.button("🚀 스케줄 자동 생성하기", use_container_width=True, type="primary")

mand_edited = st.data_editor(
    st.session_state.mand_df, use_container_width=True, height=400,
    column_config=col_cfg, key="mand_editor",
)
if not mand_edited.equals(st.session_state.mand_df):
    st.session_state.mand_df = mand_edited
    st.rerun()

st.markdown("---")

if generate_clicked:
    mandatory_dict = {e: {} for e in employees}
    for lbl in date_labels:
        for e in employees:
            val = mand_edited.at[lbl, e]
            if val in (OPT_OFF, OPT_LEAVE):
                mandatory_dict[e][label_to_day[lbl]] = val

    with st.spinner("스케줄을 짜는 중입니다..."):
        sched, counts, warnings = generate_schedule(
            year, month, employees, mandatory_dict, off_target_dict,
            min_workers, close_workers_count, store_closed_days
        )

        # 화면용 매트릭스
        matrix = {}
        for e in employees:
            matrix[e] = {day_label(month, d, wd): sched[e].get(d, ("", ""))[0]
                         for d, wd in days}
        df_matrix = pd.DataFrame.from_dict(matrix, orient="index")
        df_matrix.insert(0, "직급", [emp_info[e]["rank"] for e in employees])

        stat_data = []
        for e in employees:
            c = counts[e]
            stat_data.append({
                "직급": emp_info[e]["rank"], "이름": e,
                "목표 휴무": off_target_dict[e], "실제 휴무": c['OFF'], "연차": c['LEAVE'],
                "총 쉬는 날": c['OFF'] + c['LEAVE'],
                "A(오픈)": c['A'], "B(평일마감)": c['B'], "C(주말마감)": c['C'],
            })
        df_stats = pd.DataFrame(stat_data)

        excel_data = build_excel(store_name, year, month, int(default_off),
                                 employees, emp_info, sched)

        st.session_state.res_matrix = df_matrix
        st.session_state.res_stats = df_stats
        st.session_state.res_excel = excel_data
        st.session_state.res_warn = warnings

if "res_matrix" in st.session_state:
    st.header("📊 생성된 스케줄")

    if st.session_state.res_warn:
        for w in st.session_state.res_warn:
            st.warning(w)
    else:
        st.success("✅ 조건 충돌 없이 스케줄이 완성되었습니다!")

    st.download_button(
        "📥 엑셀 파일(.xlsx)로 다운로드",
        data=st.session_state.res_excel,
        file_name=f"{store_name}_{year}-{month:02d}_스케줄표.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("※ 색상(정휴 노랑·신청 굵게)과 연차/주말 서식은 다운로드한 엑셀에 모두 반영됩니다.")

    tab1, tab2 = st.tabs(["🗓️ 스케줄표", "📊 개인별 통계"])
    with tab1:
        st.dataframe(st.session_state.res_matrix, use_container_width=True)
    with tab2:
        st.dataframe(st.session_state.res_stats, use_container_width=True, hide_index=True)
